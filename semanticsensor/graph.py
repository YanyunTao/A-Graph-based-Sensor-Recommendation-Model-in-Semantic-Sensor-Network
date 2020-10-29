# conding=utf-8
# class GraphSet:
import openpyxl
import random
import time
import csv
from numpy import *
import fast
import numpy as np
import quick_sort

def initdatagraph(inputFile):
    graphSet = []
    vertexSet = []
    edgeSet = {}
    with open(inputFile, "r", encoding='utf-8') as fin:
        lineNum = -1
        # curVertexSet =[]
        # curEdgeSet = {}
        for line in fin:
            line = line.encode('utf-8').decode('utf-8-sig')
            lineList = line.strip().split(",")
            if not lineList:
                print("Class GraphSet __init__() line split error!")
                exit()
            if lineNum >= -1 and lineList[1] != 'http://www.semanticweb.org/zhaoqian/ontologies/2019/8/untitled-ontology-8#sensor_type':
                # 字符串截取
                lineList[0] = lineList[0].split('#')[-1]
                lineList[1] = lineList[1].split('#')[-1]
                lineList[2] = lineList[2].split('#')[-1]
                currentGraph = (lineList[0], lineList[1], lineList[2])
                # 同类型的节点，直接判断节点是否存在于集合中
                if lineList[0] not in vertexSet:
                    vertexSet.append(lineList[0])
                if lineList[1] not in vertexSet:
                    vertexSet.append(lineList[1])
                # 添加边需要将每个节点有 几条边考虑在内，结构{节点：{邻接节点:数值}，节点：{邻接节点:数值}}
                if edgeSet:
                    if lineList[0] not in list(edgeSet.keys()):
                        interrect = {}
                        interrect[lineList[1]] = lineList[2]
                        edgeSet[lineList[0]] = interrect
                    elif lineList[0] in list(edgeSet.keys()) and lineList[1] not in edgeSet[lineList[0]].keys():
                        edgeSet[lineList[0]][lineList[1]] = lineList[2]
                    else:
                        continue

                else:
                    interrect = {}
                    interrect[lineList[1]] = lineList[2]
                    edgeSet[lineList[0]] = interrect
                # 将三元组填加到图中
                graphSet.append(currentGraph)
            lineNum += 1

    return vertexSet, edgeSet, graphSet



# 匹配两个list,判断是否cover，list2是否包含list1
def matchfunction(list1,list2):
    flage=True
    if len(list1)>len(list2):
        flage=False
    for i in list1:
        if i not in list2:
            flage=False
    return flage

# 设定个函数读取表格
def readtable(tablename,inode,iedgeset):
    wb = openpyxl.load_workbook(tablename)
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    colA = ws['A']  # 获取整个C列
    low=0
    k=1
    result={}
    for i in colA:
        if i.value == inode[0]:
            low=k
            break
        k=k+1
    for i in range(low,low+8):
        if ws.cell(row=i,column=2).value.split('#')[-1] in iedgeset.keys():
            result[ws.cell(row=i,column=2).value.split('#')[-1]]=[ws.cell(row=i,column=3).value,ws.cell(row=i,column=4).value,ws.cell(row=i,column=5).value]
    # 返回值格式：{属性值：[小，大，标记]}
    return result




def write_back(result,filename):
    with open(filename, "a+", encoding='utf-8', newline='') as csvfile:
        writer = csv.writer(csvfile, dialect='excel')
        for i in result:
            writer.writerow(i)


def tanh(x):
    return (np.exp(x) - np.exp(-x)) / (np.exp(x) + np.exp(-x))

#
#     # 设定一个函数用来筛选节点 包括邻居/SAW
#     # 检索图节点集合[]、边集合{}，数据图节点集合、边集合，阈值T
#     这个函数只是根据阈值将传感器删选出来，不进行排序，排序操作使用TOPSIS算法进行
# 改进：添加映射函数 将传感器经纬度的差映射到[0，1]范围内 y=1/(x+1)
# 偏好 增加映射函数 1-1/（x+1）
# 经纬度的差和传感器偏好 权重各占0.5
# reverse -- 排序规则，reverse = True 降序 ， reverse = False 升序（默认）

def filterfunction(inode,iedgeset,sedgeset,t,k):
    table=readtable('./data/table3.xlsx', inode, iedgeset)
    print(len(sedgeset))
    # 返回满足要求的节点
    result={}
    final0=[]
    final1=[]
    sum=0
    max_pre=0
    max_jw=0
    for i in iedgeset.keys():
        if i not in ['lat','lon']:
            sum=sum+1

    for i in sedgeset.keys():
        if matchfunction(iedgeset.keys(),sedgeset[i].keys()):
            preference = 0
            # lat =math.pow((iedgeset['lat']-float(sedgeset[i]['lat'])), 2)
            # lon =math.pow(iedgeset['lon']-float(sedgeset[i]['lon']),2)
            lat=abs(iedgeset['lat']-float(sedgeset[i]['lat']))
            lon=abs(iedgeset['lon']-float(sedgeset[i]['lon']))
            # j为传感器属性
            for j in iedgeset.keys():
                # 是否确认传感器属性？
                if j not in ['lat','lon']:
                    if table[j][2]==1:
                        preference=preference+iedgeset[j]*(float(sedgeset[i][j])-float(table[j][0]))/(float(table[j][1])-float(table[j][0]))
                    elif table[j][2]==-1:
                        preference=preference+iedgeset[j]*(float(table[j][1])-float(sedgeset[i][j]))/(float(table[j][1])-float(table[j][0]))
                    else:
                        continue
            # prescore=pow(preference,2)-1;
            prescore=preference/sum
            lat_lon=1 /(pow(lat+lon, 1 / 3) + 1)
            final0.append(prescore)
            final1.append(lat_lon)
            # lat_lon=1/(pow(lat+lon,2))
            # prescore=tanh(preference)

            if lat_lon > max_jw:
                max_jw = lat_lon
            if prescore>max_pre:
                max_pre=prescore
            # 数值越大越好
            # score=(1/(math.sqrt(lat+lon)+1))/(1/(math.sqrt(lat+lon)+1)+prescore)*0.1+prescore/(1/(math.sqrt(lat+lon)+1)+prescore)*0.9
            # score =1 /(pow(lat+lon, 1 / 3) + 1) + prescore
            # score=lat_lon+ prescore
            score=lat_lon/(lat_lon+prescore)*0.1+prescore/(lat_lon+prescore)*0.9
            # score=-lat-lon+preference
            # score=1/((lat+lon)+1)+ 1-1/(prescore+1)
            # score=prescore+lat_lon
            if score>t:
                # print("-----------------------------")
                # print(prescore)
                if len(result.keys())<k:
                        result[i]=score
                else:
                    B =sorted(result.items(), key=lambda item: item[1],reverse=True)
                    # print(B)
                    if score>B[k-1][1]:
                        result.pop(B[k-1][0])
                        result[i]=score
        else:
            continue
    print("最大偏好得分："+str(max_pre))
    print("最大latlon得分："+str(max_jw))
    print(score)
    final = sorted(result.items(), key=lambda item: item[1], reverse=True)
    print(final0)
    print(final1)
    return final

# 使用改进的快速非支配排序算法
# # 参数：数据集的边集合sedgeset、筛选出的传感器节点集合finalnodeset、检索的传感器偏好集合perfrenceset

def fastfunction(sedgeset,finalnodeset,perfrenceset):
    max_prefernce=['accuracy_of_measurement','sensitivity', 'life']
    min_perference=['the_response_time','the_startup_time', 'the_energy_consumption']
    input_data=[]
    result=[]
    for i in finalnodeset:
        inter=[]
        for j in perfrenceset:
            # print(sedgeset[i[0]][j])
            if j in max_prefernce:
                inter.append(float(sedgeset[i[0]][j]))
            if j in min_perference:
                inter.append(-float(sedgeset[i[0]][j]))
        input_data.append(inter)
    firstresult=fast.fast_non_dominated_sort(input_data)
    # print(input_data)
    for i in range(len(firstresult[0])):
        result.append(finalnodeset[firstresult[0][i]])
    for j in range(len(filterresult[1])):
        result.append(finalnodeset[firstresult[1][j]])
    # for k in range(len(filterresult[2])):
    #     result.append(finalnodeset[firstresult[1][k]])
    return result


def quikfast(sedgeset,finalnodeset,perfrenceset):
    max_prefernce = ['accuracy_of_measurement', 'sensitivity', 'life']
    min_perference = ['the_response_time', 'the_startup_time', 'the_energy_consumption']
    input_data = []
    result = []
    for i in finalnodeset:
        inter = []
        for j in perfrenceset:
            # print(sedgeset[i[0]][j])
            if j in max_prefernce:
                inter.append(float(sedgeset[i[0]][j]))
            if j in min_perference:
                inter.append(-float(sedgeset[i[0]][j]))
        input_data.append(inter)
    firstresult = quick_sort.mysort.get_index(quick_sort.mysort.QuickSort(input_data,0,len(input_data)-1))
    # print(input_data)
    for i in range(len(firstresult)):
        result.append(finalnodeset[firstresult[i]])
    return result



# TOPSIS算法
# 参数：数据集的边集合sedgeset、筛选出的传感器节点集合finalnodeset、检索的传感器偏好集合perfrenceset
def TOPSIS(sedgeset,finalnodeset,perfrenceset,num):
    # 创建矩阵存储每个节点对应的传感器属性值
    max_prefernce=['accuracy_of_measurement','sensitivity', 'life']
    min_perference=['the_response_time','the_startup_time', 'the_energy_consumption']
    matrix=np.zeros((num,len(perfrenceset)))
    list=[0]*len(perfrenceset)
    # print(matrix)
    # r=matrix.size

    for i in range(num):
        for j in range(len(perfrenceset)):
            matrix[i][j]=float(sedgeset[finalnodeset[i][0]][perfrenceset[j]])

#     规则化
    for i in range(len(perfrenceset)):
        for j in range(num):
            list[i]=list[i]+math.pow(matrix[j][i],2)
        list[i]=math.sqrt(list[i])


    for i in range(num):
        for j in range(len(perfrenceset)):
            if perfrenceset[j] in max_prefernce:
                matrix[i][j]=matrix[i][j]/list[j]
            else:
                matrix[i][j]=-matrix[i][j]/list[j]
    # print(matrix)
    matrix=np.array(matrix)
    # print("------------------------------")
    # print(matrix)

#     获取最优点和最差点

    max_point=[]
    min_point=[]
    for i in range(len(perfrenceset)):
        max_point.append(max(matrix[:,i]))
        min_point.append(min(matrix[:,i]))
    # for i in range(len(perfrenceset)):
    #     if perfrenceset[i] in max_prefernce:
    #         max_point.append(max(matrix[:,i]))
    #         min_point.append(min(matrix[:,i]))
    #     if perfrenceset[i] in min_perference:
    #         max_point.append(min(matrix[:,i]))
    #         min_point.append(max(matrix[:,i]))

# 计算到最优点和最差点的欧式距离
    result={}
    for i in range(num):
        m=0
        n=0
        for j in range(len(perfrenceset)):
            m=m+math.pow((matrix[i][j]-max_point[j]),2)
            n=n+math.pow((matrix[i][j]-min_point[j]),2)
        toMax=math.sqrt(m)
        toMin=math.sqrt(n)
        score=toMin/(toMax-toMin)
        result[finalnodeset[i][0]]=score
# 排序
    B = sorted(result.items(), key=lambda item: item[1], reverse=True)

    return B



# SAW算法
# 参数：数据集的边集合sedgeset、筛选出的传感器节点集合finalnodeset、检索的传感器偏好集合perfrenceset
def SAW(sedgeset,finalnodeset,perfrenceset,num,plist):
    # 创建矩阵存储每个节点对应的传感器属性值
    max_prefernce=['accuracy_of_measurement','sensitivity', 'life']
    min_perference=['the_response_time','the_startup_time', 'the_energy_consumption']
    matrix=np.zeros((num,len(perfrenceset)))
    list=[0]*len(perfrenceset)

    for i in range(num):
        for j in range(len(perfrenceset)):
            matrix[i][j]=float(sedgeset[finalnodeset[i][0]][perfrenceset[j]])

 #     获取最大值、最小值

    max_value = []
    min_value = []
    for i in range(len(perfrenceset)):
        max_value.append(max(matrix[:, i]))
        min_value.append(min(matrix[:, i]))

#     规则化
    for i in range(num):
        for j in range(len(perfrenceset)):
            if perfrenceset[j] in max_prefernce:
                if(max_value[j] == min_value[j]):
                    matrix[i][j]=0;
                else:
                    matrix[i][j] = (matrix[i][j] - min_value[j]) / (max_value[j] - min_value[j]);
            if perfrenceset[j] in min_perference:
                if(max_value[j] == min_value[j]):
                    matrix[i][j]=0;
                else:
                    matrix[i][j] = (max_value[j] - matrix[i][j]) / (max_value[j] - min_value[j]);
    # print(matrix)

    # 计算加权值
    result = {}
    for i in range(num):
        score=0;
        for j in range(len(perfrenceset)):
            score=score+plist[j]*matrix[i][j];
        result[finalnodeset[i][0]] = score
    # reverse = True 降序 ， reverse = False 升序（默认）
    B = sorted(result.items(), key=lambda item: item[1], reverse=True)

    return B







# 设置比较参数构建虚拟用户的满意数值
# 'the_response_time':m1,'accuracy_of_measurement':m2,'the_startup_time':m3,'sensitivity':m4,'life':m5,'the_energy_consumption':m6
goal=[10,0.7,150,100,5,10]
m=[1,0.6,0.8,1,0.4,1]
# goal=[30,0.7,70,100,5,10]
# m=[0.6,1,1,0.6,0.8,1]
index=[0,3,5]
indexcontext=['the_response_time','accuracy_of_measurement','the_startup_time','sensitivity','life','the_energy_consumption']

# num=300
# top-k
# k=1
# 阈值t
# t=0.95
k=3
# for i in range(50,350,50):
num=300
t = 0.5
myresult=[]
# 'the_response_time':m1,'accuracy_of_measurement':m2,'the_startup_time':m3,'sensitivity':m4,'life':m5,'the_energy_consumption':m6
for type in ['一氧化碳传感器', '二氧化氮传感器', '二氧化硫传感器','光学传感器','加速度传感器','印刷传感器','压力传感器','温度传感器','温湿度传感器','湿度传感器','磁力传感器','紫外线传感器','臭氧传感器']:
    number_correct = 0
    # 结果记录
    interresult=[]
    # print(type)
    # s初始化数据
    z, n, q = initdatagraph('./data/'+type+'.csv')
    # 记录检索时间
    starttime = time.time()
    filterresult=filterfunction([type],{'lat':-78,'lon':40,'the_response_time':m[0],'accuracy_of_measurement':m[1],'the_startup_time':m[2],'sensitivity':m[3],'life':m[4],'the_energy_consumption':m[5]},n,t,num)
    result1=quikfast(n,filterresult,indexcontext)
    print("reslt1-------------------------------------------------------")
    print(len(result1))
    # result=TOPSIS(n,result1,indexcontext,len(result1))
    result=SAW(n,result1,indexcontext,len(result1),m)

    # print(result1)
    endtime = time.time()
    dtime = endtime - starttime
    # 计算准确率
    r = 1
    max_prefernce = ['accuracy_of_measurement', 'sensitivity', 'life']
    min_perference = ['the_response_time', 'the_startup_time', 'the_energy_consumption']
    for i in result:
        internumber = 0;
        if r<=k:
            for j in index:
                inter=indexcontext[j]
                if inter in max_prefernce:
                    if float(n[i[0]][inter]) >= goal[j]:
                        internumber=internumber+1
                else:
                    if float(n[i[0]][inter]) <= goal[j]:
                        internumber=internumber+1
                if(internumber==len(index)):
                    number_correct=number_correct+1
        else:
            break
        r = r + 1

    if len(result)<k:
        correctrate=number_correct/len(result)
    else:
        correctrate=number_correct/k
    interresult.append(correctrate)
    interresult.append(dtime)
    myresult.append(interresult)
write_back(myresult,"top"+str(k)+".csv")



# print(result)
